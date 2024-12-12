import requests
import openpyxl
import time
from openpyxl import Workbook
import os

# URL to fetch data from ESP32 (update with your ESP32's IP address)
ESP32_URL = "http://192.168.1.7/data"  # Change this IP address to your ESP32's address

# Path to your Excel file
file_path = r"D:\STEM Projects\Sensor Data.xlsx"  # Update this with your file path

# Initialize Excel file if it doesn't exist
def initialize_excel():
    try:
        workbook = openpyxl.load_workbook(file_path)
        if "Sensor Data" not in workbook.sheetnames:
            sheet = workbook.create_sheet("Sensor Data")
            headers = ["Timestamp", "Temperature (°C)", "Humidity (%)", "Air Quality (PPM)", "Flame Detected"]
            sheet.append(headers)
            workbook.save(file_path)
            print("Worksheet 'Sensor Data' created successfully.")
        else:
            print("Excel file and worksheet found and loaded.")
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Sensor Data"
        headers = ["Timestamp", "Temperature (°C)", "Humidity (%)", "Air Quality (PPM)", "Flame Detected"]
        sheet.append(headers)
        workbook.save(file_path)
        print("Excel file created successfully.")

# Function to append sensor data to Excel file
def append_to_excel(data):
    try:
        workbook = openpyxl.load_workbook(file_path)
        if "Sensor Data" in workbook.sheetnames:
            sheet = workbook["Sensor Data"]
        else:
            sheet = workbook.create_sheet("Sensor Data")
            headers = ["Timestamp", "Temperature (°C)", "Humidity (%)", "Air Quality (PPM)", "Flame Detected"]
            sheet.append(headers)

        # Adding data to the next empty row
        timestamp = time.strftime("%Y-%m-%d %H:%M:%S")
        new_row = [
            timestamp,
            data.get("temperature", "N/A"),
            data.get("humidity", "N/A"),
            data.get("airQuality", "N/A"),  # Air quality data
            data.get("flameDetected", "No Flame")  # Flame detection data
        ]
        sheet.append(new_row)
        workbook.save(file_path)
        print(f"Data appended at {timestamp}: {new_row}")
    except Exception as e:
        print(f"Error appending data to Excel: {e}")

# Function to fetch sensor data from ESP32
def fetch_sensor_data():
    try:
        response = requests.get(ESP32_URL, timeout=5)
        if response.status_code == 200:
            return response.json()
        else:
            print(f"Error fetching data: Status code {response.status_code}")
    except requests.RequestException as e:
        print(f"Request error: {e}")
    return None

# Main function to run the data fetching and updating Excel
def main():
    initialize_excel()
    while True:
        data = fetch_sensor_data()
        if data:
            append_to_excel(data)
        time.sleep(5)  # Wait for 5 seconds before the next fetch

if __name__ == "__main__":
    main()
